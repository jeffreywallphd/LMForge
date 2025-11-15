# --- Combined Imports ---
import json
import logging
import re
import time
from io import BytesIO
from typing import Dict, Any
from urllib.parse import urlparse

import markdown
import pdfplumber
import requests
from bs4 import BeautifulSoup
from django.http import StreamingHttpResponse, HttpRequest, HttpResponse
from django.shortcuts import render
from django.views import View
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

# --- Model and Util Imports ---
from ..models.scraped_data import ScrapedData  # Import the model to save data
from ..utils.content_extractor import extract_article_content # From 'updates'

# --- Setup from 'updates' ---
logger = logging.getLogger(__name__)

# Constants
MAX_TITLE_LENGTH = 100  # Maximum length for ScrapedData title field
MAX_URL_TITLE_LENGTH = 95  # Maximum length when using URL as title (leave room for "scraped")

# --- Helper from 'original' ---
def remove_emojis(text):
    """
    Removes emojis and other 4-byte characters from a string.
    """
    if not text:
        return ""
    # Regex to match most common emojis, symbols, and pictographs
    emoji_pattern = re.compile(
        "["
        "\U0001F600-\U0001F64F"  # emoticons
        "\U0001F300-\U0001F5FF"  # symbols & pictographs
        "\U0001F680-\U0001F6FF"  # transport & map symbols
        "\U0001F1E0-\U0001F1FF"  # flags (iOS)
        "\u2600-\u26FF"          # miscellaneous symbols
        "\u2700-\u27BF"          # dingbats
        "]+",
        flags=re.UNICODE,
    )
    return emoji_pattern.sub(r"", text)

# --- Merged ScrapeDataView ---
# This view keeps the streaming structure from 'original'
# but uses the improved generic scraping logic from 'updates'.
class ScrapeDataView(View):
    def stream_scrape_events(self, request):
        """
        A generator function that performs scraping and yields Server-Sent Events.
        (Kept from 'original')
        """
        url = request.GET.get('url')
        # 'title' is now fetched with a default, like in 'updates'
        title = request.GET.get('title', '') 
        source_type = request.GET.get('source_type')

        def send_event(event_type, data):
            """Helper to format data as a Server-Sent Event."""
            json_data = json.dumps({"type": event_type, "data": data})
            return f"data: {json_data}\n\n"

        if not url:
            yield send_event('error', {'message': 'Please provide a URL.'})
            return

        try:
            # --- Initial Progress Update ---
            yield send_event('progress', {'message': f"Connecting to {url}..."})

            content = None
            file_type = None

            # --- Reddit Scraper (Kept from 'original') ---
            if source_type == 'reddit':
                parsed_url = urlparse(url)
                api_url = url.rstrip('/') + ".json"
                headers = {'User-Agent': 'My-Django-Scraper-App/1.2'}
                
                yield send_event('progress', {'message': f"Requesting data from Reddit API: {api_url}"})
                response = requests.get(api_url, headers=headers)
                response.raise_for_status()
                data = response.json()

                if '/comments/' in parsed_url.path: # Handle specific post
                    file_type = 'reddit_post'
                    yield send_event('progress', {'message': 'Parsing Reddit post and comments...'})
                    
                    post_data = data[0]['data']['children'][0]['data']
                    post_title = post_data.get('title', 'No Title')
                    post_author = post_data.get('author', 'Unknown Author')
                    post_text = post_data.get('selftext', 'No content text.')
                    
                    content_lines = [
                        f"Title: {post_title}",
                        f"Author: u/{post_author}",
                        "--- POST CONTENT ---",
                        post_text,
                        "\n--- COMMENTS ---"
                    ]

                    comments_data = data[1]['data']['children']
                    for comment in comments_data:
                        if 'data' in comment and 'body' in comment['data']:
                            comment_author = comment['data'].get('author', 'Unknown')
                            comment_body = comment['data'].get('body', '')
                            content_lines.append(f"\n> u/{comment_author}:\n{comment_body}\n")
                    
                    content = "\n".join(content_lines)
                    # Use Reddit post title if user didn't provide one
                    if not title.strip():
                        title = post_title

                elif parsed_url.path.startswith('/r/'): # Handle subreddit
                    file_type = 'reddit_subreddit_full'
                    subreddit_name = parsed_url.path.split('/')[2]
                    posts = data['data']['children']
                    yield send_event('progress', {'message': f'Found {len(posts)} posts in r/{subreddit_name}. Scraping each...'})
                    
                    content_lines = [f"Scraped Posts and Comments from r/{subreddit_name}:\n" + "="*40]
                    for i, post_item in enumerate(posts):
                        post_data = post_item['data']
                        post_title = post_data.get('title', 'No Title')
                        permalink = post_data.get('permalink')

                        if not permalink: continue
                        
                        yield send_event('progress', {'message': f'({i+1}/{len(posts)}) Scraping post: "{post_title}"'})
                        
                        post_api_url = f"https://www.reddit.com{permalink.rstrip('/')}.json"
                        
                        try:
                            time.sleep(0.5) # Respect Reddit's rate limits
                            post_response = requests.get(post_api_url, headers=headers)
                            post_response.raise_for_status()
                            post_and_comment_data = post_response.json()
                            
                            post_content_data = post_and_comment_data[0]['data']['children'][0]['data']
                            post_text = post_content_data.get('selftext', '')
                            if post_text:
                                content_lines.append(f"\n--- POST CONTENT ---\n{post_text}\n")

                            content_lines.append("--- COMMENTS ---")
                            comments_data = post_and_comment_data[1]['data']['children']
                            if not comments_data:
                                content_lines.append("No comments found for this post.")
                            else:
                                for comment in comments_data:
                                    if 'data' in comment and 'body' in comment['data']:
                                        comment_author = comment['data'].get('author', 'Unknown')
                                        comment_body = comment['data'].get('body', '')
                                        content_lines.append(f"\n> u/{comment_author}:\n{comment_body}\n")
                        except Exception as post_e:
                            content_lines.append(f"\n[Could not fetch content for post. Error: {str(post_e)}]")
                    
                    content = "\n".join(content_lines)
                    # Use subreddit name in title if user didn't provide one
                    if not title.strip():
                        title = f"Scrape of r/{subreddit_name}"

                else:
                    yield send_event('error', {'message': 'Invalid Reddit URL.'})
                    return
            
            # --- Generic URL Scraper (Logic from 'updates') ---
            else:
                yield send_event('progress', {'message': 'Scraping generic URL...'})
                
                try:
                    response = requests.get(url, timeout=15)
                    response.raise_for_status()
                except Exception as e:
                    logger.exception("Failed to fetch url %s", url)
                    yield send_event('error', {'message': f"Failed to fetch URL: {e}"})
                    return

                content_type: str = response.headers.get("content-type", "").lower()

                # JSON
                if "application/json" in content_type or (response.text and response.text.strip().startswith("{")):
                    yield send_event('progress', {'message': 'Parsing JSON...'})
                    content = json.dumps(response.json(), indent=2)
                    file_type = "json"

                # XML-ish
                elif any(x in content_type for x in ("application/xml", "text/xml", "application/rss+xml")):
                    yield send_event('progress', {'message': 'Parsing XML/RSS...'})
                    content = response.text
                    file_type = "xml"

                # Plain text
                elif "text/plain" in content_type:
                    yield send_event('progress', {'message': 'Parsing plain text...'})
                    content = response.text
                    file_type = "text"

                # CSV
                elif "text/csv" in content_type or url.lower().endswith(".csv"):
                    yield send_event('progress', {'message': 'Parsing CSV...'})
                    content = response.text
                    file_type = "csv"

                # XLSX / Excel
                elif any(x in content_type for x in ("excel", "spreadsheetml", "vnd.openxmlformats")) or url.lower().endswith(".xlsx"):
                    yield send_event('progress', {'message': 'Parsing Excel (XLSX)...'})
                    try:
                        bio = BytesIO(response.content)
                        wb = load_workbook(filename=bio, read_only=True)
                        ws = wb[wb.sheetnames[0]]
                        rows = []
                        for row in ws.iter_rows(values_only=True):
                            rows.append(",".join([str(c) if c is not None else "" for c in row]))
                        content = "\n".join(rows)
                        file_type = "xlsx"
                    except Exception as e:
                        logger.exception("Failed to parse xlsx: %s", e)
                        yield send_event('error', {'message': "Failed to parse xlsx file"})
                        return
                
                # HTML (using the new 'extractor' logic from 'updates')
                else:
                    yield send_event('progress', {'message': 'Parsing HTML...'})
                    try:
                        result: Dict[str, Any] = extract_article_content(response.content, url)
                        content = result.get("body", "") or ""

                        # If extractor returned an empty body, fall back to simple parsing
                        if not content.strip():
                            logger.info("Extractor returned empty body for %s; falling back to simple HTML parsing", url)
                            soup = BeautifulSoup(response.content, "html.parser")
                            article = soup.find("article") or soup.find(class_="content") or soup.find("main")
                            if article:
                                text = article.get_text("\n\n", strip=True)
                            else:
                                body = soup.body
                                text = body.get_text("\n\n", strip=True) if body else soup.get_text("\n\n", strip=True)
                            content = text

                        # If user did not provide a title, prefer the extracted one
                        if not title or not title.strip():
                            extracted_title = result.get("title") or ""
                            if extracted_title:
                                title = extracted_title # 'title' var is updated
                        file_type = "html"
                    except Exception as e:
                        logger.exception("Extractor failed, falling back to simple HTML parsing: %s", e)
                        soup = BeautifulSoup(response.content, "html.parser")
                        article = soup.find("article") or soup.find(class_="content") or soup.find("main")
                        if article:
                            text = article.get_text("\n\n", strip=True)
                        else:
                            body = soup.body
                            text = body.get_text("\n\n", strip=True) if body else soup.get_text("\n\n", strip_True)
                        content = text
                        file_type = "html"


            # --- Save to DB and send final result (Merged logic) ---
            yield send_event('progress', {'message': 'Scraping complete. Saving to database...'})
            
            # Use emoji remover from 'original'
            cleaned_content = remove_emojis(content)

            # Use title logic from 'updates'
            # Use provided title, or extracted title, or fallback to URL
            final_title = title or (url[:MAX_URL_TITLE_LENGTH] if url else "scraped")
            
            ScrapedData.objects.create(
                url=url,
                file_type=file_type,
                content=cleaned_content,
                # 'binary_content' field removed, as 'updates' logic parses XLSX to text
                title=final_title[:MAX_TITLE_LENGTH] # Ensure max length
            )

            final_data = {
                'success': f'Successfully saved data from {url} to the database.',
                'url': url,
                'file_type': file_type,
                'content': cleaned_content
            }
            yield send_event('complete', final_data)
        
        except Exception as e:
            logger.exception("An unexpected error occurred during scrape: %s", e)
            yield send_event('error', {'message': f'An unexpected error occurred: {str(e)}'})
            return

    def get(self, request):
        """
        Returns a streaming response to send live scraping updates.
        (Kept from 'original')
        """
        response = StreamingHttpResponse(self.stream_scrape_events(request), content_type='text/event-stream')
        response['Cache-Control'] = 'no-cache'
        return response

# --- UploadPDFView (Replaced with 'updates' version) ---
class UploadPDFView(APIView):
    """Upload a PDF and convert it to text/html/json as requested."""

    def post(self, request: Request) -> Response:
        pdf_file = request.FILES.get("pdf_file")
        output_format = request.POST.get("output_format") or request.data.get("output_format") or "text"
        title = request.POST.get("title") or request.data.get("title") or (getattr(pdf_file, 'name', '') if pdf_file else "uploaded_pdf")

        if not pdf_file:
            return Response({"error": "No PDF file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            text_parts = []
            with pdfplumber.open(pdf_file) as pdf:
                for p in pdf.pages:
                    text_parts.append(p.extract_text() or "")
            text = "\n\n".join([t for t in text_parts if t])

            if output_format == "html":
                content = markdown.markdown(text)
                file_type = "html"
            elif output_format == "json":
                content = json.dumps({"text": text}, indent=2)
                file_type = "json"
            else:
                content = text
                file_type = "text"

            scraped_record: ScrapedData = ScrapedData.objects.create(
                url="uploaded_pdf", # Use placeholder URL
                file_type=file_type,
                content=content,
                pdf_file=pdf_file, # Save the original file (assuming model has this field)
                title=title[:MAX_TITLE_LENGTH]
            )
        except Exception as e:
            logger.exception("PDF conversion failed: %s", e)
            return Response({"error": "Failed to convert PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            "success": "PDF converted and saved", 
            "id": scraped_record.id, 
            "file_type": scraped_record.file_type, 
            "content": scraped_record.content
        }, status=status.HTTP_200_OK)

# --- scrape_view (Replaced with 'updates' version) ---
def scrape_view(request: HttpRequest) -> HttpResponse:
    """Render the scrape view with the latest scraped data."""
    latest: ScrapedData | None = ScrapedData.objects.order_by("-created_at").first()
    # 'updates' file used 'scraped' as context key, 'original' used 'latest_scraped_data'
    # I'll use 'latest_scraped_data' to match the 'original' template context.
    return render(request, "scrape.html", {"latest_scraped_data": latest})


# --- SaveManualTextView (Replaced with 'updates' version) ---
class SaveManualTextView(APIView):
    """Save manually entered text to ScrapedData."""
    
    def post(self, request: Request) -> Response:
        text: str | None = request.data.get("text") or request.POST.get("text")
        title: str = request.data.get("title") or request.POST.get("title") or "manual"
        if not text:
            return Response({"error": "No text provided"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            scraped_record: ScrapedData = ScrapedData.objects.create(
                url="manual", # Use placeholder URL
                file_type="text",
                content=text,
                title=title[:MAX_TITLE_LENGTH]
            )
        except Exception as e:
            logger.exception("Failed to save manual text: %s", e)
            return Response({"error": "Failed to save text"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # The 'original' file returned the latest data. The 'updates' did not.
        # I will add it back to match the original's expected response.
        return Response({
            "success": "Text saved", 
            "id": scraped_record.id, 
            "file_type": scraped_record.file_type,
            "content": scraped_record.content # Send back the content
        }, status=status.HTTP_200_OK)